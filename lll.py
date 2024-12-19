import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import sqlite3
import os
from datetime import datetime, timedelta
import calendar
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# --- Настройка базы данных ---
def initialize_db():
    conn = sqlite3.connect('class_journal.db')
    cursor = conn.cursor()
    
    # Таблица классов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT NOT NULL
        )
    ''')

    # Таблица предметов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_name TEXT NOT NULL
        )
    ''')

    # Таблица учеников
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER,
            full_name TEXT NOT NULL,
            birth_date TEXT,
            gender TEXT,
            FOREIGN KEY(class_id) REFERENCES classes(id)
        )
    ''')

    # Таблица оценок
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS grades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            subject_id INTEGER,
            grade TEXT,
            comment TEXT,
            date TEXT,
            FOREIGN KEY(student_id) REFERENCES students(id),
            FOREIGN KEY(subject_id) REFERENCES subjects(id)
        )
    ''')

    # Таблица пропусков
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS absences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            reason TEXT,
            date TEXT,
            FOREIGN KEY(student_id) REFERENCES students(id)
        )
    ''')

    # Таблица замечаний
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS remarks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            remark TEXT,
            status TEXT,
            date TEXT,
            FOREIGN KEY(student_id) REFERENCES students(id)
        )
    ''')

    # Таблица домашних заданий
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS homework (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER,
            subject_id INTEGER,
            lesson_date TEXT,
            description TEXT,
            due_date TEXT,
            FOREIGN KEY(class_id) REFERENCES classes(id),
            FOREIGN KEY(subject_id) REFERENCES subjects(id)
        )
    ''')

    # Таблица выполнения домашних заданий
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS homework_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            homework_id INTEGER,
            student_id INTEGER,
            file_name TEXT,
            status TEXT,
            comment TEXT,
            submission_date TEXT,
            FOREIGN KEY(homework_id) REFERENCES homework(id),
            FOREIGN KEY(student_id) REFERENCES students(id)
        )
    ''')

    # Таблица профиля пользователя
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_profile (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            patronymic TEXT,
            date_of_birth TEXT,
            gender TEXT,
            phone TEXT,
            additional_phone TEXT,
            address TEXT,
            avatar_path TEXT
        )
    ''')

    # Таблица событий для календаря
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            start_time TEXT,
            end_time TEXT,
            status TEXT,
            comments TEXT
        )
    ''')

    # Добавление некоторых предметов по умолчанию, если таблица пуста
    cursor.execute("SELECT COUNT(*) FROM subjects")
    if cursor.fetchone()[0] == 0:
        subjects = ["Математика", "Русский язык", "Физика", "Химия", "Биология", "История", "География", "Литература"]
        for subj in subjects:
            cursor.execute("INSERT INTO subjects (subject_name) VALUES (?)", (subj,))

    # Добавление некоторых классов по умолчанию, если таблица пуста
    cursor.execute("SELECT COUNT(*) FROM classes")
    if cursor.fetchone()[0] == 0:
        classes = ["10-А", "10-Б", "11-А", "11-Б"]
        for cls in classes:
            cursor.execute("INSERT INTO classes (class_name) VALUES (?)", (cls,))

    # Проверка наличия профиля пользователя, если нет - создать пустой профиль
    cursor.execute("SELECT COUNT(*) FROM user_profile")
    if cursor.fetchone()[0] == 0:
        cursor.execute("""
            INSERT INTO user_profile (first_name, last_name, patronymic, date_of_birth, gender, phone, additional_phone, address, avatar_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Имя", "Фамилия", "Отчество", "01.01.2000", "М", "1234567890", "", "Адрес", ""))
    
    conn.commit()
    conn.close()

# --- Классы для управления событиями ---
class EventDetailWindow(tk.Toplevel):
    def __init__(self, master, event):
        super().__init__(master)
        self.title("Детали события")
        self.geometry("400x300")
        self.resizable(False, False)
        self.event = event
        
        # Настройка сетки
        for i in range(7):
            self.grid_rowconfigure(i, pad=5)
        self.grid_columnconfigure(1, weight=1)
        
        # Создание меток и полей
        tk.Label(self, text="Название:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky='e')
        tk.Label(self, text=self.event['title'], font=("Arial", 10)).grid(row=0, column=1, sticky='w')
        
        tk.Label(self, text="Описание:", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky='ne')
        description_text = tk.Text(self, width=30, height=5, font=("Arial", 10))
        description_text.grid(row=1, column=1, sticky='w')
        description_text.insert('1.0', self.event['description'])
        description_text.config(state='disabled')

        tk.Label(self, text="Начало:", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky='e')
        tk.Label(self, text=self.event['start_time'].strftime("%Y-%m-%d %H:%M"), font=("Arial", 10)).grid(row=2, column=1, sticky='w')

        tk.Label(self, text="Окончание:", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky='e')
        tk.Label(self, text=self.event['end_time'].strftime("%Y-%m-%d %H:%M"), font=("Arial", 10)).grid(row=3, column=1, sticky='w')

        tk.Label(self, text="Статус:", font=("Arial", 10, "bold")).grid(row=4, column=0, sticky='e')
        tk.Label(self, text=self.event['status'], font=("Arial", 10)).grid(row=4, column=1, sticky='w')

        tk.Label(self, text="Комментарии:", font=("Arial", 10, "bold")).grid(row=5, column=0, sticky='ne')
        comments_text = tk.Text(self, width=30, height=5, font=("Arial", 10))
        comments_text.grid(row=5, column=1, sticky='w')
        comments_text.insert('1.0', self.event['comments'])
        comments_text.config(state='disabled')
        
        tk.Button(self, text="Закрыть", command=self.destroy).grid(row=6, column=1, pady=10, sticky='e')

# --- Классы Форм ---
class StudentForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, student_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.student_id = student_id
        self.title("Добавить ученика" if not student_id else "Редактировать ученика")
        self.geometry("400x500")
        
        self.create_widgets()
        if student_id:
            self.load_student_data()
    
    def create_widgets(self):
        tk.Label(self, text="ФИО").pack(pady=5)
        self.full_name_entry = tk.Entry(self)
        self.full_name_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата рождения (ДД.ММ.ГГГГ)").pack(pady=5)
        self.birth_date_entry = tk.Entry(self)
        self.birth_date_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Пол").pack(pady=5)
        self.gender_var = tk.StringVar()
        gender_combo = ttk.Combobox(self, textvariable=self.gender_var, values=["М", "Ж"], state="readonly")
        gender_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Класс").pack(pady=5)
        self.class_var = tk.StringVar()
        self.class_combo = ttk.Combobox(self, textvariable=self.class_var)
        self.load_classes()
        self.class_combo.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_student).pack(pady=20)
    
    def load_classes(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT class_name FROM classes")
        classes = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.class_combo['values'] = classes
    
    def load_student_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT full_name, birth_date, gender, class_id FROM students WHERE id=?", (self.student_id,))
        student = cursor.fetchone()
        conn.close()
        if student:
            self.full_name_entry.insert(0, student[0])
            self.birth_date_entry.insert(0, student[1])
            self.gender_var.set(student[2])
            
            # Получение названия класса
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("SELECT class_name FROM classes WHERE id=?", (student[3],))
            class_name = cursor.fetchone()
            conn.close()
            if class_name:
                self.class_var.set(class_name[0])
    
    def save_student(self):
        full_name = self.full_name_entry.get()
        birth_date = self.birth_date_entry.get()
        gender = self.gender_var.get()
        class_name = self.class_var.get()
        
        if not all([full_name, birth_date, gender, class_name]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля.")
            return
        
        # Валидация даты
        try:
            datetime.strptime(birth_date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты рождения.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        
        # Получение id класса
        cursor.execute("SELECT id FROM classes WHERE class_name=?", (class_name,))
        class_row = cursor.fetchone()
        if not class_row:
            # Если класса нет, добавить новый
            cursor.execute("INSERT INTO classes (class_name) VALUES (?)", (class_name,))
            class_id = cursor.lastrowid
        else:
            class_id = class_row[0]
        
        if self.student_id:
            cursor.execute("""
                UPDATE students 
                SET full_name=?, birth_date=?, gender=?, class_id=?
                WHERE id=?
            """, (full_name, birth_date, gender, class_id, self.student_id))
        else:
            cursor.execute("""
                INSERT INTO students (full_name, birth_date, gender, class_id)
                VALUES (?, ?, ?, ?)
            """, (full_name, birth_date, gender, class_id))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Данные ученика сохранены успешно.")
        self.destroy()

class GradeForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, add_event_callback, grade_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.add_event_callback = add_event_callback
        self.grade_id = grade_id
        self.title("Добавить оценку" if not grade_id else "Редактировать оценку")
        self.geometry("400x500")
        
        self.create_widgets()
        if grade_id:
            self.load_grade_data()
    
    def create_widgets(self):
        tk.Label(self, text="Ученик").pack(pady=5)
        self.student_var = tk.StringVar()
        self.student_combo = ttk.Combobox(self, textvariable=self.student_var)
        self.load_students()
        self.student_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Предмет").pack(pady=5)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(self, textvariable=self.subject_var)
        self.load_subjects()
        self.subject_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата (ДД.ММ.ГГГГ)").pack(pady=5)
        self.date_entry = tk.Entry(self)
        self.date_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Оценка").pack(pady=5)
        self.grade_var = tk.StringVar()
        self.grade_combo = ttk.Combobox(self, textvariable=self.grade_var, values=["5", "4", "3", "2", "1", "Н"], state="readonly")
        self.grade_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Комментарий").pack(pady=5)
        self.comment_entry = tk.Entry(self)
        self.comment_entry.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_grade).pack(pady=20)
    
    def load_students(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT full_name FROM students")
        students = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.student_combo['values'] = students
    
    def load_subjects(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT subject_name FROM subjects")
        subjects = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.subject_combo['values'] = subjects
    
    def load_grade_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT students.full_name, subjects.subject_name, grades.date, grades.grade, grades.comment 
            FROM grades 
            JOIN students ON grades.student_id = students.id
            JOIN subjects ON grades.subject_id = subjects.id
            WHERE grades.id=?
        """, (self.grade_id,))
        grade = cursor.fetchone()
        conn.close()
        if grade:
            self.student_var.set(grade[0])
            self.subject_var.set(grade[1])
            self.date_entry.insert(0, grade[2])
            self.grade_var.set(grade[3])
            self.comment_entry.insert(0, grade[4])
    
    def save_grade(self):
        student_name = self.student_var.get()
        subject_name = self.subject_var.get()
        date = self.date_entry.get()
        grade = self.grade_var.get()
        comment = self.comment_entry.get()
        
        if not all([student_name, subject_name, date, grade]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
            return
        
        # Валидация даты
        try:
            datetime.strptime(date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        
        # Получение id ученика
        cursor.execute("SELECT id FROM students WHERE full_name=?", (student_name,))
        student = cursor.fetchone()
        if not student:
            messagebox.showerror("Ошибка", "Ученик не найден.")
            conn.close()
            return
        
        # Получение id предмета
        cursor.execute("SELECT id FROM subjects WHERE subject_name=?", (subject_name,))
        subject = cursor.fetchone()
        if not subject:
            messagebox.showerror("Ошибка", "Предмет не найден.")
            conn.close()
            return
        
        if self.grade_id:
            cursor.execute("""
                UPDATE grades
                SET student_id=?, subject_id=?, date=?, grade=?, comment=?
                WHERE id=?
            """, (student[0], subject[0], date, grade, comment, self.grade_id))
            # Обновление события в календаре
            self.update_calendar_event('grade', self.grade_id, title=f"Оценка: {grade} по {subject_name}", 
                                       start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                       end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        else:
            cursor.execute("""
                INSERT INTO grades (student_id, subject_id, date, grade, comment)
                VALUES (?, ?, ?, ?, ?)
            """, (student[0], subject[0], date, grade, comment))
            grade_id = cursor.lastrowid
            # Добавление события в календарь
            self.add_calendar_event('grade', grade_id, title=f"Оценка: {grade} по {subject_name}", 
                                    start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                    end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Оценка сохранена успешно.")
        self.destroy()

    def add_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO events (title, description, start_time, end_time, status, comments)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (title, '', start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments))
        conn.commit()
        conn.close()

    def update_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Предположим, что связь между оценкой и событием происходит через title и date
        cursor.execute("""
            UPDATE events
            SET title=?, start_time=?, end_time=?, status=?, comments=?
            WHERE title LIKE ? AND start_time LIKE ?
        """, (title, start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments, f"Оценка: {self.grade_var.get()}%", start_time.strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

class AbsenceForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, add_event_callback, absence_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.add_event_callback = add_event_callback
        self.absence_id = absence_id
        self.title("Добавить пропуск" if not absence_id else "Редактировать пропуск")
        self.geometry("400x400")
        
        self.create_widgets()
        if absence_id:
            self.load_absence_data()
    
    def create_widgets(self):
        tk.Label(self, text="Ученик").pack(pady=5)
        self.student_var = tk.StringVar()
        self.student_combo = ttk.Combobox(self, textvariable=self.student_var)
        self.load_students()
        self.student_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Причина").pack(pady=5)
        self.reason_entry = tk.Entry(self)
        self.reason_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата (ДД.ММ.ГГГГ)").pack(pady=5)
        self.date_entry = tk.Entry(self)
        self.date_entry.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_absence).pack(pady=20)
    
    def load_students(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT full_name FROM students")
        students = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.student_combo['values'] = students
    
    def load_absence_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT students.full_name, absences.reason, absences.date 
            FROM absences 
            JOIN students ON absences.student_id = students.id
            WHERE absences.id=?
        """, (self.absence_id,))
        absence = cursor.fetchone()
        conn.close()
        if absence:
            self.student_var.set(absence[0])
            self.reason_entry.insert(0, absence[1])
            self.date_entry.insert(0, absence[2])
    
    def save_absence(self):
        student_name = self.student_var.get()
        reason = self.reason_entry.get()
        date = self.date_entry.get()
        
        if not all([student_name, reason, date]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля.")
            return
        
        # Валидация даты
        try:
            datetime.strptime(date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM students WHERE full_name=?", (student_name,))
        student = cursor.fetchone()
        if not student:
            messagebox.showerror("Ошибка", "Ученик не найден.")
            conn.close()
            return
        
        if self.absence_id:
            cursor.execute("""
                UPDATE absences
                SET student_id=?, reason=?, date=?
                WHERE id=?
            """, (student[0], reason, date, self.absence_id))
            # Обновление события в календаре
            self.update_calendar_event('absence', self.absence_id, title=f"Пропуск: {reason}", 
                                       start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                       end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        else:
            cursor.execute("""
                INSERT INTO absences (student_id, reason, date)
                VALUES (?, ?, ?)
            """, (student[0], reason, date))
            absence_id = cursor.lastrowid
            # Добавление события в календарь
            self.add_calendar_event('absence', absence_id, title=f"Пропуск: {reason}", 
                                    start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                    end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Пропуск сохранен успешно.")
        self.destroy()
    
    def add_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO events (title, description, start_time, end_time, status, comments)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (title, '', start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments))
        conn.commit()
        conn.close()

    def update_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Предположим, что связь между пропуском и событием происходит через title и date
        cursor.execute("""
            UPDATE events
            SET title=?, start_time=?, end_time=?, status=?, comments=?
            WHERE title LIKE ? AND start_time LIKE ?
        """, (title, start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments, f"Пропуск: {reason}%", start_time.strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

class RemarkForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, add_event_callback, remark_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.add_event_callback = add_event_callback
        self.remark_id = remark_id
        self.title("Добавить замечание" if not remark_id else "Редактировать замечание")
        self.geometry("400x500")
        
        self.create_widgets()
        if remark_id:
            self.load_remark_data()
    
    def create_widgets(self):
        tk.Label(self, text="Ученик").pack(pady=5)
        self.student_var = tk.StringVar()
        self.student_combo = ttk.Combobox(self, textvariable=self.student_var)
        self.load_students()
        self.student_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Замечание").pack(pady=5)
        self.remark_entry = tk.Text(self, height=4)
        self.remark_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Статус").pack(pady=5)
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(self, textvariable=self.status_var, values=["Не рассмотрено", "Рассмотрено"], state="readonly")
        self.status_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата (ДД.ММ.ГГГГ)").pack(pady=5)
        self.date_entry = tk.Entry(self)
        self.date_entry.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_remark).pack(pady=20)
    
    def load_students(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT full_name FROM students")
        students = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.student_combo['values'] = students
    
    def load_remark_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT students.full_name, remarks.remark, remarks.status, remarks.date 
            FROM remarks 
            JOIN students ON remarks.student_id = students.id
            WHERE remarks.id=?
        """, (self.remark_id,))
        remark = cursor.fetchone()
        conn.close()
        if remark:
            self.student_var.set(remark[0])
            self.remark_entry.insert(tk.END, remark[1])
            self.status_var.set(remark[2])
            self.date_entry.insert(0, remark[3])
    
    def save_remark(self):
        student_name = self.student_var.get()
        remark = self.remark_entry.get("1.0", tk.END).strip()
        status = self.status_var.get()
        date = self.date_entry.get()
        
        if not all([student_name, remark, status, date]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
            return
        
        # Валидация даты
        try:
            datetime.strptime(date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM students WHERE full_name=?", (student_name,))
        student = cursor.fetchone()
        if not student:
            messagebox.showerror("Ошибка", "Ученик не найден.")
            conn.close()
            return
        
        if self.remark_id:
            cursor.execute("""
                UPDATE remarks
                SET student_id=?, remark=?, status=?, date=?
                WHERE id=?
            """, (student[0], remark, status, date, self.remark_id))
            # Обновление события в календаре
            self.update_calendar_event('remark', self.remark_id, title=f"Замечание: {status}", 
                                       start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                       end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        else:
            cursor.execute("""
                INSERT INTO remarks (student_id, remark, status, date)
                VALUES (?, ?, ?, ?)
            """, (student[0], remark, status, date))
            remark_id = cursor.lastrowid
            # Добавление события в календарь
            self.add_calendar_event('remark', remark_id, title=f"Замечание: {status}", 
                                    start_time=datetime.strptime(date, "%d.%m.%Y"), 
                                    end_time=datetime.strptime(date, "%d.%m.%Y") + timedelta(hours=1))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Замечание сохранено успешно.")
        self.destroy()
    
    def add_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO events (title, description, start_time, end_time, status, comments)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (title, '', start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments))
        conn.commit()
        conn.close()

    def update_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Предположим, что связь между замечанием и событием происходит через title и date
        cursor.execute("""
            UPDATE events
            SET title=?, start_time=?, end_time=?, status=?, comments=?
            WHERE title LIKE ? AND start_time LIKE ?
        """, (title, start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments, f"Замечание: {status}%", start_time.strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

class HomeworkForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, add_event_callback, homework_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.add_event_callback = add_event_callback
        self.homework_id = homework_id
        self.title("Добавить задание" if not homework_id else "Редактировать задание")
        self.geometry("500x600")
        
        self.create_widgets()
        if homework_id:
            self.load_homework_data()
    
    def create_widgets(self):
        tk.Label(self, text="Класс").pack(pady=5)
        self.class_var = tk.StringVar()
        self.class_combo = ttk.Combobox(self, textvariable=self.class_var)
        self.load_classes()
        self.class_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Предмет").pack(pady=5)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(self, textvariable=self.subject_var)
        self.load_subjects()
        self.subject_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата урока (ДД.ММ.ГГГГ)").pack(pady=5)
        self.lesson_date_entry = tk.Entry(self)
        self.lesson_date_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Описание задания").pack(pady=5)
        self.description_text = tk.Text(self, height=10)
        self.description_text.pack(fill=tk.BOTH, padx=10, pady=5)
        
        tk.Label(self, text="Срок выполнения (ДД.ММ.ГГГГ)").pack(pady=5)
        self.due_date_entry = tk.Entry(self)
        self.due_date_entry.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_homework).pack(pady=20)
    
    def load_classes(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT class_name FROM classes")
        classes = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.class_combo['values'] = classes
    
    def load_subjects(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT subject_name FROM subjects")
        subjects = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.subject_combo['values'] = subjects
    
    def load_homework_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT classes.class_name, subjects.subject_name, homework.lesson_date, homework.description, homework.due_date
            FROM homework
            LEFT JOIN classes ON homework.class_id = classes.id
            LEFT JOIN subjects ON homework.subject_id = subjects.id
            WHERE homework.id=?
        """, (self.homework_id,))
        homework = cursor.fetchone()
        conn.close()
        if homework:
            self.class_var.set(homework[0])
            self.subject_var.set(homework[1])
            self.lesson_date_entry.insert(0, homework[2])
            self.description_text.insert(tk.END, homework[3])
            self.due_date_entry.insert(0, homework[4])
    
    def save_homework(self):
        cls = self.class_var.get()
        subject = self.subject_var.get()
        lesson_date = self.lesson_date_entry.get()
        description = self.description_text.get("1.0", tk.END).strip()
        due_date = self.due_date_entry.get()
        
        if not all([cls, subject, lesson_date, description, due_date]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
            return
        
        # Валидация дат
        try:
            datetime.strptime(lesson_date, "%d.%m.%Y")
            datetime.strptime(due_date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        
        # Получение id класса
        cursor.execute("SELECT id FROM classes WHERE class_name=?", (cls,))
        class_row = cursor.fetchone()
        if not class_row:
            messagebox.showerror("Ошибка", "Класс не найден.")
            conn.close()
            return
        class_id = class_row[0]
        
        # Получение id предмета
        cursor.execute("SELECT id FROM subjects WHERE subject_name=?", (subject,))
        subject_row = cursor.fetchone()
        if not subject_row:
            messagebox.showerror("Ошибка", "Предмет не найден.")
            conn.close()
            return
        subject_id = subject_row[0]
        
        if self.homework_id:
            cursor.execute("""
                UPDATE homework
                SET class_id=?, subject_id=?, lesson_date=?, description=?, due_date=?
                WHERE id=?
            """, (class_id, subject_id, lesson_date, description, due_date, self.homework_id))
            # Обновление события в календаре
            self.update_calendar_event('homework', self.homework_id, title=f"Домашнее задание по {subject}", 
                                       start_time=datetime.strptime(due_date, "%d.%m.%Y"), 
                                       end_time=datetime.strptime(due_date, "%d.%m.%Y") + timedelta(hours=1))
        else:
            cursor.execute("""
                INSERT INTO homework (class_id, subject_id, lesson_date, description, due_date)
                VALUES (?, ?, ?, ?, ?)
            """, (class_id, subject_id, lesson_date, description, due_date))
            homework_id = cursor.lastrowid
            # Добавление события в календарь
            self.add_calendar_event('homework', homework_id, title=f"Домашнее задание по {subject}", 
                                    start_time=datetime.strptime(due_date, "%d.%m.%Y"), 
                                    end_time=datetime.strptime(due_date, "%d.%m.%Y") + timedelta(hours=1))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Домашнее задание сохранено успешно.")
        self.destroy()
    
    def add_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO events (title, description, start_time, end_time, status, comments)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (title, '', start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments))
        conn.commit()
        conn.close()

    def update_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Предположим, что связь между домашним заданием и событием происходит через title и date
        cursor.execute("""
            UPDATE events
            SET title=?, start_time=?, end_time=?, status=?, comments=?
            WHERE title LIKE ? AND start_time LIKE ?
        """, (title, start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments, f"Домашнее задание по {title.split(' по ')[1]}%", start_time.strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

# --- Класс для отображения деталей события ---
# Уже определен выше как EventDetailWindow

# --- Класс Календаря ---
class CalendarApp(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Календарь")
        self.geometry("1100x700")
        self.resizable(True, True)

        # Параметры дневного вида
        self.start_hour = 8
        self.end_hour = 23
        self.row_height = 25  # 30 мин = 25px
        self.current_view = "month"
        self.current_date = datetime.now()

        # Верхняя панель (кнопка Сегодня + переключатели)
        self.topbar = tk.Frame(self)
        self.topbar.pack(side=tk.TOP, fill=tk.X)

        self.btn_today = tk.Button(self.topbar, text="Сегодня", command=self.go_today)
        self.btn_today.pack(side=tk.LEFT, padx=5, pady=5)

        # Кнопки переключения видов
        self.btn_month = tk.Button(self.topbar, text="Месяц", command=lambda: self.switch_view("month"))
        self.btn_month.pack(side=tk.RIGHT, padx=5, pady=5)

        self.btn_week = tk.Button(self.topbar, text="Неделя", command=lambda: self.switch_view("week"))
        self.btn_week.pack(side=tk.RIGHT, padx=5, pady=5)

        self.btn_day = tk.Button(self.topbar, text="День", command=lambda: self.switch_view("day"))
        self.btn_day.pack(side=tk.RIGHT, padx=5, pady=5)

        # Заголовок с датой и стрелками
        self.header_frame = tk.Frame(self)
        self.header_frame.pack(side=tk.TOP, fill=tk.X)

        self.nav_frame = tk.Frame(self.header_frame)
        self.nav_frame.pack(side=tk.TOP, pady=10)

        self.btn_prev = tk.Button(self.nav_frame, text="←", command=self.prev_period)
        self.btn_prev.pack(side=tk.LEFT, padx=5)

        self.header_label = tk.Label(self.nav_frame, font=("Arial", 16, "bold"))
        self.header_label.pack(side=tk.LEFT, padx=10)

        self.btn_next = tk.Button(self.nav_frame, text="→", command=self.next_period)
        self.btn_next.pack(side=tk.LEFT, padx=5)

        # Область для календаря со скроллом
        self.scroll_frame = tk.Frame(self)
        self.scroll_frame.pack(expand=True, fill=tk.BOTH)

        self.canvas = tk.Canvas(self.scroll_frame, bg="white")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scroll_y = tk.Scrollbar(self.scroll_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.scroll_y.set)

        # Фрейм контента внутри canvas
        self.content_frame = tk.Frame(self.canvas, bg="white")
        self.canvas_window = self.canvas.create_window((0,0), window=self.content_frame, anchor="nw")
        self.content_frame.bind("<Configure>", self.on_frame_configure)

        # Привязка колеса мыши для прокрутки
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

        self.day_boxes = []
        self.draw_view()

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _bind_mousewheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)  # Для Linux
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)  # Для Linux

    def _unbind_mousewheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        # Для Windows и MacOS
        if event.delta:
            self.canvas.yview_scroll(int(-event.delta/120), "units")
        # Для Linux
        elif event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")

    def switch_view(self, view):
        self.current_view = view
        self.draw_view()

    def go_today(self):
        self.current_date = datetime.now()
        self.draw_view()

    def prev_period(self):
        if self.current_view == "day":
            self.current_date -= timedelta(days=1)
        elif self.current_view == "week":
            self.current_date -= timedelta(weeks=1)
        elif self.current_view == "month":
            year = self.current_date.year
            month = self.current_date.month - 1
            if month < 1:
                month = 12
                year -= 1
            day = min(self.current_date.day, calendar.monthrange(year, month)[1])
            self.current_date = self.current_date.replace(year=year, month=month, day=day)
        self.draw_view()

    def next_period(self):
        if self.current_view == "day":
            self.current_date += timedelta(days=1)
        elif self.current_view == "week":
            self.current_date += timedelta(weeks=1)
        elif self.current_view == "month":
            year = self.current_date.year
            month = self.current_date.month + 1
            if month > 12:
                month = 1
                year += 1
            day = min(self.current_date.day, calendar.monthrange(year, month)[1])
            self.current_date = self.current_date.replace(year=year, month=month, day=day)
        self.draw_view()

    def update_header(self):
        if self.current_view == "day":
            date_str = self.current_date.strftime("%d %B %Y г.")
            self.header_label.config(text=date_str)
        elif self.current_view == "week":
            start_of_week = self.current_date - timedelta(days=self.current_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            date_str = f"Неделя: {start_of_week.strftime('%d %B %Y')} - {end_of_week.strftime('%d %B %Y')}"
            self.header_label.config(text=date_str)
        elif self.current_view == "month":
            date_str = self.current_date.strftime("%B %Y г.")
            self.header_label.config(text=date_str)
        else:
            self.header_label.config(text="")

    def update_view_buttons(self):
        default_bg = self.btn_day.cget("bg")
        default_fg = self.btn_day.cget("fg")
        self.btn_day.config(bg=default_bg, fg=default_fg)
        self.btn_week.config(bg=default_bg, fg=default_fg)
        self.btn_month.config(bg=default_bg, fg=default_fg)

        if self.current_view == "day":
            self.btn_day.config(bg="green", fg="white")
        elif self.current_view == "week":
            self.btn_week.config(bg="green", fg="white")
        elif self.current_view == "month":
            self.btn_month.config(bg="green", fg="white")

    def draw_view(self):
        # Очистим content_frame
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        self.day_boxes.clear()
        self.update_header()
        self.update_view_buttons()

        if self.current_view == "day":
            self.draw_day_view()
            self.place_day_events()
        elif self.current_view == "week":
            self.draw_week_view()
            self.place_week_events()
        elif self.current_view == "month":
            self.draw_month_view()
            self.place_month_events()
        else:
            self.canvas.create_text(200, 100, text="Вид не поддерживается", font=("Arial", 14))

        self.canvas.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    # ------------------- День -------------------
    def draw_day_view(self):
        # Дневной вид с 8:00 до 23:00
        left_margin = 80
        top_margin = 20
        timeline_canvas = tk.Canvas(self.content_frame, bg="white", highlightthickness=0)
        timeline_canvas.pack(fill=tk.BOTH, expand=True)

        total_hours = self.end_hour - self.start_hour + 1
        total_height = total_hours * self.row_height * 2 + top_margin * 2
        timeline_canvas.config(height=total_height, width=1000)

        # Подсветка сегодняшнего дня
        if self.current_date.date() == datetime.now().date():
            timeline_canvas.create_rectangle(0, 0, 2000, 2000, fill="#fffacd", outline="")

        for hour in range(self.start_hour, self.end_hour + 1):
            y = top_margin + (hour - self.start_hour) * self.row_height * 2
            timeline_canvas.create_line(left_margin, y, 2000, y, width=2, fill="black")
            time_str = f"{hour}:00"
            timeline_canvas.create_text(left_margin - 10, y, text=time_str, anchor='e', font=("Arial", 10))

            if hour < self.end_hour:
                half_y = y + self.row_height
                timeline_canvas.create_line(left_margin, half_y, 2000, half_y, fill="gray", dash=(2, 2))
                half_str = f"{hour}:30"
                timeline_canvas.create_text(left_margin - 10, half_y, text=half_str, anchor='e', font=("Arial", 8))

    def place_day_events(self):
        day_events = self.fetch_events(start_date=self.current_date.date(), end_date=self.current_date.date())
        day_events.sort(key=lambda ev: ev['start_time'])
        if not day_events:
            return

        # Найдём Canvas (он единственный в дневном режиме)
        timeline_canvas = self.content_frame.winfo_children()[0]

        left_margin = 80
        top_margin = 20

        for event in day_events:
            start_hour = event['start_time'].hour
            start_minute = event['start_time'].minute
            end_hour = event['end_time'].hour
            end_minute = event['end_time'].minute

            # Вычисляем позицию по Y с учётом минут
            start_y = top_margin + (start_hour - self.start_hour) * self.row_height * 2 + (self.row_height if start_minute >= 30 else 0)
            end_y = top_margin + (end_hour - self.start_hour) * self.row_height * 2 + (self.row_height if end_minute >= 30 else 0)

            x1 = left_margin + 10
            x2 = x1 + 200
            y1 = start_y + 5
            y2 = max(y1 + 20, end_y - 5)
            fill_color = "lightgreen" if event['status'] == 'planned' else "#d3d3d3"

            rect = timeline_canvas.create_rectangle(x1, y1, x2, y2, fill=fill_color, outline="black")
            event_text = f"{event['title']}\n{event['start_time'].strftime('%H:%M')}-{event['end_time'].strftime('%H:%M')}"
            txt = timeline_canvas.create_text(x1 + 5, (y1 + y2) / 2, text=event_text, anchor='w', font=("Arial", 10), width=190)

            def on_event_click(e, ev=event):
                self.open_event_detail(ev)

            timeline_canvas.tag_bind(rect, "<Button-1>", on_event_click)
            timeline_canvas.tag_bind(txt, "<Button-1>", on_event_click)

    # ------------------- Неделя -------------------
    def draw_week_view(self):
        days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

        week_canvas = tk.Canvas(self.content_frame, bg="white", highlightthickness=0)
        week_canvas.pack(fill=tk.BOTH, expand=True)

        col_width = 120
        row_height = self.row_height
        left_margin = 50
        top_margin = 20

        current_day_of_week = self.current_date.weekday()
        start_of_week = self.current_date - timedelta(days=current_day_of_week)
        total_hours = self.end_hour - self.start_hour
        canvas_height = (total_hours * row_height) + (row_height) + top_margin + 200
        week_canvas.config(height=canvas_height, width=1000)

        for i, day_name in enumerate(days_of_week):
            x = left_margin + i * col_width
            y = top_margin
            day_date = start_of_week + timedelta(days=i)
            if day_date.date() == datetime.now().date():
                week_canvas.create_rectangle(x, y, x + col_width, y + row_height, fill="#fffacd", outline="black")
                week_canvas.create_text(x + col_width / 2, y + row_height / 2, text=day_name, font=("Arial", 10, "bold"))
            else:
                week_canvas.create_rectangle(x, y, x + col_width, y + row_height, fill="#f0f0f0", outline="black")
                week_canvas.create_text(x + col_width / 2, y + row_height / 2, text=day_name, font=("Arial", 10))

        base_y = top_margin + row_height
        for hour in range(self.start_hour, self.end_hour + 1):
            y = base_y + (hour - self.start_hour) * row_height
            week_canvas.create_line(left_margin, y, left_margin + col_width * 7, y, fill="gray")
            time_str = f"{hour}:00"
            week_canvas.create_text(left_margin - 30, y, text=time_str, anchor='e', font=("Arial", 8))

        for i in range(8):
            x = left_margin + i * col_width
            week_canvas.create_line(x, top_margin + row_height, x, top_margin + row_height + total_hours * row_height, fill="gray")

    def place_week_events(self):
        week_start_date = self.current_date - timedelta(days=self.current_date.weekday())
        week_end_date = week_start_date + timedelta(days=6)
        week_events = self.fetch_events(start_date=week_start_date.date(),
                                        end_date=week_end_date.date())
        week_events.sort(key=lambda ev: ev['start_time'])
        if not week_events:
            return

        week_canvas = self.content_frame.winfo_children()[0]

        left_margin = 50
        top_margin = 20
        col_width = 120
        row_height = self.row_height

        base_y = top_margin + row_height
        for event in week_events:
            day_index = (event['start_time'].date() - week_start_date.date()).days
            if day_index < 0 or day_index > 6:
                continue  # Игнорируем события вне текущей недели

            start_hour = event['start_time'].hour
            start_minute = event['start_time'].minute
            end_hour = event['end_time'].hour
            end_minute = event['end_time'].minute

            # Вычисляем позицию по Y с учётом минут
            start_y = base_y + (start_hour - self.start_hour) * row_height + (row_height // 2 if start_minute >= 30 else 0)
            end_y = base_y + (end_hour - self.start_hour) * row_height + (row_height // 2 if end_minute >= 30 else 0)

            y1 = start_y + 5
            y2 = max(y1 + 20, end_y - 5)

            x = left_margin + day_index * col_width
            fill_color = "lightgreen" if event['status'] == 'planned' else "#d3d3d3"
            rect = week_canvas.create_rectangle(x + 2, y1, x + col_width - 2, y2, fill=fill_color, outline="black")
            event_text = f"{event['title']} ({event['start_time'].strftime('%H:%M')}-{event['end_time'].strftime('%H:%M')})"
            txt = week_canvas.create_text(x + col_width / 2, (y1 + y2) / 2, text=event_text, font=("Arial", 9), width=col_width - 10)

            def on_event_click(e, ev=event):
                self.open_event_detail(ev)
            
            week_canvas.tag_bind(rect, "<Button-1>", on_event_click)
            week_canvas.tag_bind(txt, "<Button-1>", on_event_click)

    # ------------------- Месяц -------------------
    def draw_month_view(self):
        self.day_boxes = []
        days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        year = self.current_date.year
        month = self.current_date.month
        first_day_weekday, num_days = calendar.monthrange(year, month)

        month_canvas = tk.Canvas(self.content_frame, bg="white", highlightthickness=0)
        month_canvas.pack(fill=tk.BOTH, expand=True)
        cell_width = 150
        cell_height = 100
        top_margin = 40
        left_margin = 50

        # Шапка дней недели
        for i, d in enumerate(days_of_week):
            x = left_margin + i * cell_width
            y = top_margin
            month_canvas.create_rectangle(x, y, x + cell_width, y + (cell_height * 0.5), fill="#e0e0e0", outline="black")
            month_canvas.create_text(
                x + cell_width / 2, 
                y + (cell_height * 0.5) / 2, 
                text=d, 
                font=("Arial", 10, "bold")
            )

        start_date = datetime(year, month, 1)
        # Корректируем первый день недели для календаря (0=Пн, 6=Вс)
        calendar_start_date = start_date - timedelta(days=start_date.weekday())

        for week in range(6):  # Обычно 6 недель в месячном виде
            for wday in range(7):
                day_date = calendar_start_date + timedelta(days=week * 7 + wday)
                x = left_margin + wday * cell_width
                y = top_margin + cell_height * 0.5 + week * cell_height
                month_canvas.create_rectangle(x, y, x + cell_width, y + cell_height, fill="white", outline="black")

                if day_date.month == month:
                    text_color = "black"
                else:
                    text_color = "gray"

                # Подсветка сегодняшнего дня
                if day_date.date() == datetime.now().date() and day_date.month == month:
                    month_canvas.create_rectangle(
                        x, y, x + cell_width, y + cell_height, 
                        fill="#fffacd", outline="black"
                    )

                # Отображение номера дня
                month_canvas.create_text(
                    x + 5, y + 10, 
                    text=str(day_date.day), 
                    anchor='nw', 
                    fill=text_color, 
                    font=("Arial", 10)
                )

                # Сохранение информации о ячейке дня для последующего использования
                self.day_boxes.append((day_date, x, y, cell_width, cell_height))

    def place_month_events(self):
        if not self.day_boxes:
            return

        min_date = self.day_boxes[0][0].date()
        max_date = self.day_boxes[-1][0].date()
        month_events = self.fetch_events(start_date=min_date, end_date=max_date)
        events_by_day = {}
        for e in month_events:
            d = e['start_time'].date()
            events_by_day.setdefault(d, []).append(e)

        month_canvas = self.content_frame.winfo_children()[0]

        for day_date, x, y, w, h in self.day_boxes:
            day_events = events_by_day.get(day_date, [])
            day_events.sort(key=lambda ev: ev['start_time'])
            displayed_events = day_events[:2]
            extra_count = len(day_events) - 2 if len(day_events) > 2 else 0

            event_y = y + 25
            for ev in displayed_events:
                start_str = ev['start_time'].strftime("%H:%M")
                end_str = ev['end_time'].strftime("%H:%M")
                event_str = f"{start_str}-{end_str} {ev['title']}"
                fill_color = "lightgreen" if ev['status'] == 'planned' else "#d3d3d3"
                rect = month_canvas.create_rectangle(x + 5, event_y, x + w - 5, event_y + 20, fill=fill_color, outline="black")
                txt = month_canvas.create_text(x + 10, event_y + 10, text=event_str, anchor='w', font=("Arial", 9))
    
                def on_event_click(e, ev=ev):
                    self.open_event_detail(ev)
    
                month_canvas.tag_bind(rect, "<Button-1>", on_event_click)
                month_canvas.tag_bind(txt, "<Button-1>", on_event_click)
                event_y += 25

            # Обработчик клика на ячейку дня или на "ещё..."
            def on_day_click_empty(e, d=day_date):
                self.current_date = datetime(d.year, d.month, d.day)
                self.switch_view("day")
    
            # Создание прозрачного прямоугольника для захвата кликов
            transparent_rect = month_canvas.create_rectangle(x + 1, y + 1, x + w - 1, y + h - 1, outline="", fill="", width=0)
            month_canvas.tag_bind(transparent_rect, "<Button-1>", on_day_click_empty)

            if extra_count > 0:
                more_str = f"ещё {extra_count}..."
                more_txt = month_canvas.create_text(
                    x + 10, event_y + 10, 
                    text=more_str, 
                    anchor='w', 
                    font=("Arial", 9, "underline"), 
                    fill="blue"
                )
                # Привязка клика к тексту "ещё..."
                month_canvas.tag_bind(more_txt, "<Button-1>", on_day_click_empty)

    def open_event_detail(self, event):
        # Метод для открытия окна деталей события
        EventDetailWindow(self, event)

    def fetch_events(self, start_date, end_date):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT title, description, start_time, end_time, status, comments
            FROM events
            WHERE date(start_time) BETWEEN ? AND ?
            ORDER BY start_time
        """, (start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))
        rows = cursor.fetchall()
        conn.close()
        events = []
        for row in rows:
            events.append({
                'title': row[0],
                'description': row[1],
                'start_time': datetime.strptime(row[2], "%Y-%m-%d %H:%M"),
                'end_time': datetime.strptime(row[3], "%Y-%m-%d %H:%M"),
                'status': row[4],
                'comments': row[5]
            })
        return events

    # --- Экспорт и Печать ---
    class PrintJournalWindow(tk.Toplevel):
        def __init__(self, master):
            super().__init__(master)
            self.title("Печать журнала")
            self.geometry("400x300")
            
            self.create_widgets()
        
        def create_widgets(self):
            tk.Label(self, text="Выберите раздел для печати:", font=("Arial", 12)).pack(pady=20)
            
            sections = ["Лист всеобуча", "Оценки", "Пропуски", "Замечания", "Домашнее задание", "Календарь"]
            self.section_var = tk.StringVar()
            self.section_combo = ttk.Combobox(self, textvariable=self.section_var, values=sections, state="readonly")
            self.section_combo.pack(pady=10)
            
            tk.Button(self, text="Сохранить как PDF", command=self.save_pdf).pack(pady=20)
        
        def save_pdf(self):
            section = self.section_var.get()
            if not section:
                messagebox.showerror("Ошибка", "Пожалуйста, выберите раздел для печати.")
                return
            
            file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                     filetypes=[("PDF files", "*.pdf")])
            if file_path:
                c = canvas.Canvas(file_path, pagesize=A4)
                width, height = A4
                c.setFont("Helvetica", 16)
                c.drawString(100, height - 50, f"Электронный классный журнал - {section}")
                c.setFont("Helvetica", 12)
                
                # Получение данных из выбранного раздела
                data = []
                headers = []
                if section == "Лист всеобуча":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT students.id, students.full_name, students.birth_date, students.gender, classes.class_name
                        FROM students
                        LEFT JOIN classes ON students.class_id = classes.id
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["ID", "ФИО", "Дата рождения", "Пол", "Класс"]
                elif section == "Оценки":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT grades.id, students.full_name, subjects.subject_name, grades.grade, grades.comment, grades.date 
                        FROM grades 
                        JOIN students ON grades.student_id = students.id
                        JOIN subjects ON grades.subject_id = subjects.id
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["ID", "Ученик", "Предмет", "Оценка", "Комментарий", "Дата"]
                elif section == "Пропуски":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT absences.id, students.full_name, absences.reason, absences.date, classes.class_name
                        FROM absences
                        LEFT JOIN students ON absences.student_id = students.id
                        LEFT JOIN classes ON students.class_id = classes.id
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["ID", "Ученик", "Причина", "Дата", "Класс"]
                elif section == "Замечания":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT remarks.id, students.full_name, remarks.remark, remarks.status, remarks.date, classes.class_name
                        FROM remarks
                        LEFT JOIN students ON remarks.student_id = students.id
                        LEFT JOIN classes ON students.class_id = classes.id
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["ID", "Ученик", "Замечание", "Статус", "Дата", "Класс"]
                elif section == "Домашнее задание":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT homework.id, classes.class_name, subjects.subject_name, homework.lesson_date, 
                               homework.description, homework.due_date
                        FROM homework
                        LEFT JOIN classes ON homework.class_id = classes.id
                        LEFT JOIN subjects ON homework.subject_id = subjects.id
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["ID", "Класс", "Предмет", "Дата урока", "Описание", "Срок выполнения"]
                elif section == "Календарь":
                    conn = sqlite3.connect('class_journal.db')
                    cursor = conn.cursor()
                    cursor.execute("""
                        SELECT title, description, start_time, end_time, status, comments
                        FROM events
                        ORDER BY start_time
                    """)
                    data = cursor.fetchall()
                    conn.close()
                    headers = ["Название", "Описание", "Начало", "Окончание", "Статус", "Комментарии"]
                else:
                    messagebox.showerror("Ошибка", "Неверный раздел.")
                    return
                
                # Вывод данных в PDF
                y = height - 80
                x_offsets = [50, 150, 250, 350, 450, 550, 650]  # Примерные смещения по X
                for idx, header in enumerate(headers):
                    if idx < len(x_offsets):
                        c.drawString(x_offsets[idx], y, header)
                y -= 20
                for row in data:
                    for idx, item in enumerate(row):
                        if idx < len(x_offsets):
                            text = str(item)
                            if len(text) > 20:
                                text = text[:17] + "..."
                            c.drawString(x_offsets[idx], y, text)
                    y -= 20
                    if y < 50:
                        c.showPage()
                        y = height - 50
                c.save()
                messagebox.showinfo("Печать", f"Журнал сохранен как {os.path.basename(file_path)}")

# --- Классы для управления студентами, оценками, пропусками, замечаниями, домашними заданиями ---
class StudentListWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Лист всеобуча")
        self.geometry("800x600")
        
        self.create_widgets()
        self.load_students()
    
    def create_widgets(self):
        # Поиск
        search_frame = tk.Frame(self)
        search_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(search_frame, text="Поиск по ФИО:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(search_frame, text="Поиск", command=self.search_students).pack(side=tk.LEFT)
        tk.Button(search_frame, text="Сброс", command=self.load_students).pack(side=tk.LEFT, padx=5)
        
        # Таблица
        columns = ("ID", "ФИО", "Дата рождения", "Пол", "Класс")
        self.tree = ttk.Treeview(self, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # Кнопки
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Добавить ученика", command=self.add_student).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Редактировать ученика", command=self.edit_student).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить ученика", command=self.delete_student).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_students(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT students.id, students.full_name, students.birth_date, students.gender, classes.class_name 
            FROM students
            LEFT JOIN classes ON students.class_id = classes.id
        """)
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def search_students(self):
        query = self.search_var.get()
        if not query:
            self.load_students()
            return
        
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT students.id, students.full_name, students.birth_date, students.gender, classes.class_name 
            FROM students
            LEFT JOIN classes ON students.class_id = classes.id
            WHERE students.full_name LIKE ?
        """, ('%' + query + '%',))
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def add_student(self):
        StudentForm(self, self.load_students)
    
    def edit_student(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите ученика для редактирования.")
            return
        values = self.tree.item(selected, 'values')
        StudentForm(self, self.load_students, student_id=values[0])
    
    def delete_student(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите ученика для удаления.")
            return
        values = self.tree.item(selected, 'values')
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить ученика {values[1]}?")
        if confirm:
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM students WHERE id=?", (values[0],))
            conn.commit()
            conn.close()
            self.load_students()
            messagebox.showinfo("Успех", "Ученик удален успешно.")
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Лист всеобуча"
            
            # Заголовки
            headers = ["ID", "ФИО", "Дата рождения", "Пол", "Класс"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Данные
            for row_num, row_id in enumerate(self.tree.get_children(), 2):
                row = self.tree.item(row_id)['values']
                for col_num, cell_value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно экспортированы в {os.path.basename(file_path)}")

class GradesWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Оценки")
        self.geometry("900x600")
        
        self.create_widgets()
        self.load_grades()
    
    def create_widgets(self):
        # Поиск
        search_frame = tk.Frame(self)
        search_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(search_frame, text="Поиск по ученику:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(search_frame, text="Поиск", command=self.search_grades).pack(side=tk.LEFT)
        tk.Button(search_frame, text="Сброс", command=self.load_grades).pack(side=tk.LEFT, padx=5)
        
        # Таблица
        columns = ("ID", "Ученик", "Предмет", "Оценка", "Комментарий", "Дата")
        self.tree = ttk.Treeview(self, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # Кнопки
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Добавить оценку", command=self.add_grade).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Редактировать оценку", command=self.edit_grade).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить оценку", command=self.delete_grade).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_grades(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT grades.id, students.full_name, subjects.subject_name, grades.grade, grades.comment, grades.date 
            FROM grades 
            JOIN students ON grades.student_id = students.id
            JOIN subjects ON grades.subject_id = subjects.id
        """)
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def search_grades(self):
        query = self.search_var.get()
        if not query:
            self.load_grades()
            return
        
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT grades.id, students.full_name, subjects.subject_name, grades.grade, grades.comment, grades.date 
            FROM grades 
            JOIN students ON grades.student_id = students.id
            JOIN subjects ON grades.subject_id = subjects.id
            WHERE students.full_name LIKE ?
        """, ('%' + query + '%',))
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def add_grade(self):
        GradeForm(self, self.load_grades, self.add_event)
    
    def edit_grade(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите оценку для редактирования.")
            return
        values = self.tree.item(selected, 'values')
        GradeForm(self, self.load_grades, self.add_event, grade_id=values[0])
    
    def delete_grade(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите оценку для удаления.")
            return
        values = self.tree.item(selected, 'values')
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить оценку ID {values[0]}?")
        if confirm:
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM grades WHERE id=?", (values[0],))
            conn.commit()
            conn.close()
            self.load_grades()
            messagebox.showinfo("Успех", "Оценка удалена успешно.")
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Оценки"
            
            # Заголовки
            headers = ["ID", "Ученик", "Предмет", "Оценка", "Комментарий", "Дата"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Данные
            for row_num, row_id in enumerate(self.tree.get_children(), 2):
                row = self.tree.item(row_id)['values']
                for col_num, cell_value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно экспортированы в {os.path.basename(file_path)}")

class AbsencesWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Пропуски")
        self.geometry("800x600")
        
        self.create_widgets()
        self.load_absences()
    
    def create_widgets(self):
        # Поиск
        search_frame = tk.Frame(self)
        search_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(search_frame, text="Поиск по ученику:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(search_frame, text="Поиск", command=self.search_absences).pack(side=tk.LEFT)
        tk.Button(search_frame, text="Сброс", command=self.load_absences).pack(side=tk.LEFT, padx=5)
        
        # Таблица
        columns = ("ID", "Ученик", "Причина", "Дата", "Класс")
        self.tree = ttk.Treeview(self, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # Кнопки
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Добавить пропуск", command=self.add_absence).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Редактировать пропуск", command=self.edit_absence).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить пропуск", command=self.delete_absence).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_absences(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT absences.id, students.full_name, absences.reason, absences.date, classes.class_name
            FROM absences
            LEFT JOIN students ON absences.student_id = students.id
            LEFT JOIN classes ON students.class_id = classes.id
        """)
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def search_absences(self):
        query = self.search_var.get()
        if not query:
            self.load_absences()
            return
        
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT absences.id, students.full_name, absences.reason, absences.date, classes.class_name
            FROM absences
            LEFT JOIN students ON absences.student_id = students.id
            LEFT JOIN classes ON students.class_id = classes.id
            WHERE students.full_name LIKE ?
        """, ('%' + query + '%',))
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def add_absence(self):
        AbsenceForm(self, self.load_absences, self.add_event)
    
    def edit_absence(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите пропуск для редактирования.")
            return
        values = self.tree.item(selected, 'values')
        AbsenceForm(self, self.load_absences, self.add_event, absence_id=values[0])
    
    def delete_absence(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите пропуск для удаления.")
            return
        values = self.tree.item(selected, 'values')
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить пропуск ID {values[0]}?")
        if confirm:
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM absences WHERE id=?", (values[0],))
            conn.commit()
            conn.close()
            self.load_absences()
            messagebox.showinfo("Успех", "Пропуск удален успешно.")
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Пропуски"
            
            # Заголовки
            headers = ["ID", "Ученик", "Причина", "Дата", "Класс"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Данные
            for row_num, row_id in enumerate(self.tree.get_children(), 2):
                row = self.tree.item(row_id)['values']
                for col_num, cell_value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно экспортированы в {os.path.basename(file_path)}")

class RemarksWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Замечания")
        self.geometry("900x600")
        
        self.create_widgets()
        self.load_remarks()
    
    def create_widgets(self):
        # Поиск
        search_frame = tk.Frame(self)
        search_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(search_frame, text="Поиск по ученику:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        tk.Button(search_frame, text="Поиск", command=self.search_remarks).pack(side=tk.LEFT)
        tk.Button(search_frame, text="Сброс", command=self.load_remarks).pack(side=tk.LEFT, padx=5)
        
        # Таблица
        columns = ("ID", "Ученик", "Замечание", "Статус", "Дата", "Класс")
        self.tree = ttk.Treeview(self, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # Кнопки
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Добавить замечание", command=self.add_remark).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Редактировать замечание", command=self.edit_remark).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить замечание", command=self.delete_remark).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_remarks(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT remarks.id, students.full_name, remarks.remark, remarks.status, remarks.date, classes.class_name
            FROM remarks
            LEFT JOIN students ON remarks.student_id = students.id
            LEFT JOIN classes ON students.class_id = classes.id
        """)
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def search_remarks(self):
        query = self.search_var.get()
        if not query:
            self.load_remarks()
            return
        
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT remarks.id, students.full_name, remarks.remark, remarks.status, remarks.date, classes.class_name
            FROM remarks
            LEFT JOIN students ON remarks.student_id = students.id
            LEFT JOIN classes ON students.class_id = classes.id
            WHERE students.full_name LIKE ?
        """, ('%' + query + '%',))
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def add_remark(self):
        RemarkForm(self, self.load_remarks, self.add_event)
    
    def edit_remark(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите замечание для редактирования.")
            return
        values = self.tree.item(selected, 'values')
        RemarkForm(self, self.load_remarks, self.add_event, remark_id=values[0])
    
    def delete_remark(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите замечание для удаления.")
            return
        values = self.tree.item(selected, 'values')
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить замечание ID {values[0]}?")
        if confirm:
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM remarks WHERE id=?", (values[0],))
            conn.commit()
            conn.close()
            self.load_remarks()
            messagebox.showinfo("Успех", "Замечание удалено успешно.")
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Замечания"
            
            # Заголовки
            headers = ["ID", "Ученик", "Замечание", "Статус", "Дата", "Класс"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Данные
            for row_num, row_id in enumerate(self.tree.get_children(), 2):
                row = self.tree.item(row_id)['values']
                for col_num, cell_value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно экспортированы в {os.path.basename(file_path)}")

class SubmissionForm(tk.Toplevel):
    def __init__(self, master, homework_id, add_event_callback):
        super().__init__(master)
        self.homework_id = homework_id
        self.add_event_callback = add_event_callback
        self.title("Отправить выполнение")
        self.geometry("400x300")
        
        self.create_widgets()
    
    def create_widgets(self):
        tk.Label(self, text="Ученик").pack(pady=10)
        self.student_var = tk.StringVar()
        self.student_combo = ttk.Combobox(self, textvariable=self.student_var)
        self.load_students()
        self.student_combo.pack(fill=tk.X, padx=20)
        
        tk.Label(self, text="Файл выполнения").pack(pady=10)
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(self, textvariable=self.file_path_var, state='readonly')
        self.file_entry.pack(fill=tk.X, padx=20)
        tk.Button(self, text="Выбрать файл", command=self.select_file).pack(pady=5)
        
        tk.Button(self, text="Отправить выполнение", command=self.submit_homework).pack(pady=20)
    
    def load_students(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Получение класса из домашнего задания
        cursor.execute("""
            SELECT classes.class_name 
            FROM homework
            JOIN classes ON homework.class_id = classes.id
            WHERE homework.id=?
        """, (self.homework_id,))
        class_row = cursor.fetchone()
        if class_row:
            class_name = class_row[0]
            cursor.execute("""
                SELECT full_name FROM students
                WHERE class_id = (
                    SELECT class_id FROM homework WHERE id=?
                )
            """, (self.homework_id,))
            students = [row[0] for row in cursor.fetchall()]
        else:
            students = []
        conn.close()
        self.student_combo['values'] = students
    
    def select_file(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.file_path_var.set(file_path)
    
    def submit_homework(self):
        student_name = self.student_var.get()
        file_path = self.file_path_var.get()
        
        if not all([student_name, file_path]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля.")
            return
        
        # Проверка существования файла
        if not os.path.isfile(file_path):
            messagebox.showerror("Ошибка", "Файл не найден.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM students WHERE full_name=?", (student_name,))
        student = cursor.fetchone()
        if not student:
            messagebox.showerror("Ошибка", "Ученик не найден.")
            conn.close()
            return
        student_id = student[0]
        
        # Сохранение файла в папку 'submissions'
        submissions_dir = "submissions"
        if not os.path.exists(submissions_dir):
            os.makedirs(submissions_dir)
        file_name = os.path.basename(file_path)
        destination = os.path.join(submissions_dir, file_name)
        try:
            with open(file_path, 'rb') as src, open(destination, 'wb') as dst:
                dst.write(src.read())
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")
            conn.close()
            return
        
        submission_date = datetime.now().strftime("%d.%m.%Y")
        status = "Выполнено"
        comment = ""
        
        # Проверка, существует ли уже запись о выполнении
        cursor.execute("""
            SELECT id FROM homework_submissions
            WHERE homework_id=? AND student_id=?
        """, (self.homework_id, student_id))
        submission = cursor.fetchone()
        if submission:
            # Обновление существующей записи
            cursor.execute("""
                UPDATE homework_submissions
                SET file_name=?, status=?, comment=?, submission_date=?
                WHERE id=?
            """, (file_name, status, comment, submission_date, submission[0]))
        else:
            # Вставка новой записи
            cursor.execute("""
                INSERT INTO homework_submissions (homework_id, student_id, file_name, status, comment, submission_date)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (self.homework_id, student_id, file_name, status, comment, submission_date))
        
        conn.commit()
        conn.close()
        
        messagebox.showinfo("Успех", "Выполнение домашнего задания отправлено успешно.")
        self.destroy()

# --- Класс Профиля Пользователя ---
class ProfileWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Профиль пользователя")
        self.geometry("500x600")
        self.resizable(False, False)
        
        self.avatar_size = (150, 150)
        self.avatar_image = None  # Для хранения объекта изображения
        
        self.create_widgets()
        self.load_profile_data()
    
    def create_widgets(self):
        # Раздел "Общие данные"
        general_frame = tk.LabelFrame(self, text="Общие данные", padx=10, pady=10)
        general_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        # Фамилия
        tk.Label(general_frame, text="Фамилия:").grid(row=0, column=0, sticky='e', pady=5)
        self.last_name_var = tk.StringVar()
        self.last_name_entry = tk.Entry(general_frame, textvariable=self.last_name_var)
        self.last_name_entry.grid(row=0, column=1, pady=5, sticky='w')
        
        # Имя
        tk.Label(general_frame, text="Имя:").grid(row=1, column=0, sticky='e', pady=5)
        self.first_name_var = tk.StringVar()
        self.first_name_entry = tk.Entry(general_frame, textvariable=self.first_name_var)
        self.first_name_entry.grid(row=1, column=1, pady=5, sticky='w')
        
        # Отчество
        tk.Label(general_frame, text="Отчество:").grid(row=2, column=0, sticky='e', pady=5)
        self.patronymic_var = tk.StringVar()
        self.patronymic_entry = tk.Entry(general_frame, textvariable=self.patronymic_var)
        self.patronymic_entry.grid(row=2, column=1, pady=5, sticky='w')
        
        # Дата рождения
        tk.Label(general_frame, text="Дата рождения (ДД.ММ.ГГГГ):").grid(row=3, column=0, sticky='e', pady=5)
        self.dob_var = tk.StringVar()
        self.dob_entry = tk.Entry(general_frame, textvariable=self.dob_var)
        self.dob_entry.grid(row=3, column=1, pady=5, sticky='w')
        
        # Пол
        tk.Label(general_frame, text="Пол:").grid(row=4, column=0, sticky='e', pady=5)
        self.gender_var = tk.StringVar()
        self.gender_combo = ttk.Combobox(general_frame, textvariable=self.gender_var, values=["М", "Ж"], state="readonly")
        self.gender_combo.grid(row=4, column=1, pady=5, sticky='w')
        
        # Контактный телефон
        tk.Label(general_frame, text="Контактный телефон:").grid(row=5, column=0, sticky='e', pady=5)
        self.phone_var = tk.StringVar()
        self.phone_entry = tk.Entry(general_frame, textvariable=self.phone_var)
        self.phone_entry.grid(row=5, column=1, pady=5, sticky='w')
        
        # Дополнительный контактный телефон
        tk.Label(general_frame, text="Дополнительный телефон:").grid(row=6, column=0, sticky='e', pady=5)
        self.additional_phone_var = tk.StringVar()
        self.additional_phone_entry = tk.Entry(general_frame, textvariable=self.additional_phone_var)
        self.additional_phone_entry.grid(row=6, column=1, pady=5, sticky='w')
        
        # Адрес
        tk.Label(general_frame, text="Адрес:").grid(row=7, column=0, sticky='ne', pady=5)
        self.address_text = tk.Text(general_frame, width=30, height=3)
        self.address_text.grid(row=7, column=1, pady=5, sticky='w')
        
        # Раздел "Аватарка"
        avatar_frame = tk.LabelFrame(self, text="Аватарка", padx=10, pady=10)
        avatar_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        self.avatar_label = tk.Label(avatar_frame, text="Нет аватарки", width=20, height=10, bg="#f0f0f0")
        self.avatar_label.pack(pady=10)
        
        tk.Button(avatar_frame, text="Загрузить фото", command=self.upload_avatar).pack(pady=5)
        
        # Кнопки "Сохранить изменения" и "Очистить поля"
        button_frame = tk.Frame(self)
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Сохранить изменения", command=self.save_profile).pack(side=tk.LEFT, padx=10)
        tk.Button(button_frame, text="Очистить поля", command=self.clear_fields).pack(side=tk.LEFT, padx=10)
    
    def load_profile_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT first_name, last_name, patronymic, date_of_birth, gender, phone, additional_phone, address, avatar_path FROM user_profile WHERE id=1")
        profile = cursor.fetchone()
        conn.close()
        if profile:
            self.first_name_var.set(profile[0])
            self.last_name_var.set(profile[1])
            self.patronymic_var.set(profile[2])
            self.dob_var.set(profile[3])
            self.gender_var.set(profile[4])
            self.phone_var.set(profile[5])
            self.additional_phone_var.set(profile[6])
            self.address_text.insert('1.0', profile[7])
            avatar_path = profile[8]
            if avatar_path and os.path.isfile(avatar_path):
                self.load_avatar_image(avatar_path)
    
    def load_avatar_image(self, path):
        try:
            image = Image.open(path)
            image = image.resize(self.avatar_size, Image.ANTIALIAS)
            self.avatar_image = ImageTk.PhotoImage(image)
            self.avatar_label.config(image=self.avatar_image, text="")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить аватарку: {e}")
            self.avatar_label.config(text="Нет аватарки", image='')
    
    def upload_avatar(self):
        file_path = filedialog.askopenfilename(
            title="Выберите изображение",
            filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")]
        )
        if file_path:
            # Проверка формата
            if not file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
                messagebox.showerror("Ошибка", "Неподдерживаемый формат изображения. Выберите PNG или JPEG.")
                return
            # Проверка размера и изменение
            try:
                image = Image.open(file_path)
                image.thumbnail(self.avatar_size, Image.ANTIALIAS)
                # Сохранение изображения в папку avatars
                avatars_dir = "avatars"
                if not os.path.exists(avatars_dir):
                    os.makedirs(avatars_dir)
                # Генерация уникального имени файла, чтобы избежать перезаписи
                unique_id = datetime.now().strftime("%Y%m%d%H%M%S")
                ext = os.path.splitext(file_path)[1]
                filename = f"user_avatar_{unique_id}{ext}"
                save_path = os.path.join(avatars_dir, filename)
                image.save(save_path)
                self.load_avatar_image(save_path)
                self.avatar_path = save_path
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось обработать изображение: {e}")
    
    def save_profile(self):
        first_name = self.first_name_var.get().strip()
        last_name = self.last_name_var.get().strip()
        patronymic = self.patronymic_var.get().strip()
        date_of_birth = self.dob_var.get().strip()
        gender = self.gender_var.get().strip()
        phone = self.phone_var.get().strip()
        additional_phone = self.additional_phone_var.get().strip()
        address = self.address_text.get("1.0", tk.END).strip()
        avatar_path = getattr(self, 'avatar_path', '')
        
        # Валидация обязательных полей
        if not all([first_name, last_name, date_of_birth, gender, phone, address]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
            return
        
        # Валидация даты рождения
        try:
            datetime.strptime(date_of_birth, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты рождения. Используйте ДД.ММ.ГГГГ.")
            return
        
        # Сохранение данных в базу
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE user_profile
            SET first_name=?, last_name=?, patronymic=?, date_of_birth=?, gender=?, phone=?, additional_phone=?, address=?, avatar_path=?
            WHERE id=1
        """, (first_name, last_name, patronymic, date_of_birth, gender, phone, additional_phone, address, avatar_path))
        conn.commit()
        conn.close()
        
        messagebox.showinfo("Успех", "Данные профиля успешно сохранены.")
    
    def clear_fields(self):
        self.first_name_var.set("")
        self.last_name_var.set("")
        self.patronymic_var.set("")
        self.dob_var.set("")
        self.gender_var.set("")
        self.phone_var.set("")
        self.additional_phone_var.set("")
        self.address_text.delete("1.0", tk.END)
        self.avatar_label.config(image='', text="Нет аватарки")
        self.avatar_image = None
        self.avatar_path = ""

# --- Класс Домашнего Задания ---
class HomeworkWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Домашнее задание")
        self.geometry("1000x600")
        
        self.create_widgets()
        self.load_homework()
    
    def create_widgets(self):
        # Поиск и фильтрация
        filter_frame = tk.Frame(self)
        filter_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(filter_frame, text="Класс:").pack(side=tk.LEFT)
        self.class_var = tk.StringVar()
        self.class_combo = ttk.Combobox(filter_frame, textvariable=self.class_var)
        self.load_classes()
        self.class_combo.pack(side=tk.LEFT, padx=5)
        
        tk.Label(filter_frame, text="Предмет:").pack(side=tk.LEFT)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(filter_frame, textvariable=self.subject_var)
        self.load_subjects()
        self.subject_combo.pack(side=tk.LEFT, padx=5)
        
        tk.Label(filter_frame, text="Дата урока (ДД.ММ.ГГГГ):").pack(side=tk.LEFT)
        self.lesson_date_var = tk.StringVar()
        self.lesson_date_entry = tk.Entry(filter_frame, textvariable=self.lesson_date_var)
        self.lesson_date_entry.pack(side=tk.LEFT, padx=5)
        
        tk.Button(filter_frame, text="Фильтр", command=self.filter_homework).pack(side=tk.LEFT, padx=5)
        tk.Button(filter_frame, text="Сброс", command=self.load_homework).pack(side=tk.LEFT, padx=5)
        
        # Таблица
        columns = ("ID", "Класс", "Предмет", "Дата урока", "Описание", "Срок выполнения")
        self.tree = ttk.Treeview(self, columns=columns, show='headings')
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "Описание":
                self.tree.column(col, width=300)
            else:
                self.tree.column(col, width=140)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # Кнопки
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Добавить задание", command=self.add_homework).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Редактировать задание", command=self.edit_homework).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Удалить задание", command=self.delete_homework).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Отправить выполнение", command=self.submit_homework).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
    
    def load_classes(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT class_name FROM classes")
        classes = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.class_combo['values'] = classes
    
    def load_subjects(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT subject_name FROM subjects")
        subjects = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.subject_combo['values'] = subjects
    
    def load_homework(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT homework.id, classes.class_name, subjects.subject_name, homework.lesson_date, 
                   homework.description, homework.due_date
            FROM homework
            LEFT JOIN classes ON homework.class_id = classes.id
            LEFT JOIN subjects ON homework.subject_id = subjects.id
        """)
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def filter_homework(self):
        cls = self.class_var.get()
        subject = self.subject_var.get()
        lesson_date = self.lesson_date_var.get()
        
        query = """
            SELECT homework.id, classes.class_name, subjects.subject_name, homework.lesson_date, 
                   homework.description, homework.due_date
            FROM homework
            LEFT JOIN classes ON homework.class_id = classes.id
            LEFT JOIN subjects ON homework.subject_id = subjects.id
            WHERE 1=1
        """
        params = []
        
        if cls:
            query += " AND classes.class_name = ?"
            params.append(cls)
        if subject:
            query += " AND subjects.subject_name = ?"
            params.append(subject)
        if lesson_date:
            query += " AND homework.lesson_date = ?"
            params.append(lesson_date)
        
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute(query, tuple(params))
        for row in cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
        conn.close()
    
    def add_homework(self):
        HomeworkForm(self, self.load_homework, self.add_event)
    
    def edit_homework(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите задание для редактирования.")
            return
        values = self.tree.item(selected, 'values')
        HomeworkForm(self, self.load_homework, self.add_event, homework_id=values[0])
    
    def delete_homework(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите задание для удаления.")
            return
        values = self.tree.item(selected, 'values')
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить задание ID {values[0]}?")
        if confirm:
            conn = sqlite3.connect('class_journal.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM homework WHERE id=?", (values[0],))
            conn.commit()
            conn.close()
            self.load_homework()
            messagebox.showinfo("Успех", "Задание удалено успешно.")
    
    def submit_homework(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите задание для выполнения.")
            return
        values = self.tree.item(selected, 'values')
        SubmissionForm(self, values[0], self.add_event)
    
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Домашнее задание"
            
            # Заголовки
            headers = ["ID", "Класс", "Предмет", "Дата урока", "Описание", "Срок выполнения"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Данные
            for row_num, row_id in enumerate(self.tree.get_children(), 2):
                row = self.tree.item(row_id)['values']
                for col_num, cell_value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            wb.save(file_path)
            messagebox.showinfo("Экспорт", f"Данные успешно экспортированы в {os.path.basename(file_path)}")

class HomeworkForm(tk.Toplevel):
    def __init__(self, master, refresh_callback, add_event_callback, homework_id=None):
        super().__init__(master)
        self.refresh_callback = refresh_callback
        self.add_event_callback = add_event_callback
        self.homework_id = homework_id
        self.title("Добавить задание" if not homework_id else "Редактировать задание")
        self.geometry("500x600")
        
        self.create_widgets()
        if homework_id:
            self.load_homework_data()
    
    def create_widgets(self):
        tk.Label(self, text="Класс").pack(pady=5)
        self.class_var = tk.StringVar()
        self.class_combo = ttk.Combobox(self, textvariable=self.class_var)
        self.load_classes()
        self.class_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Предмет").pack(pady=5)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(self, textvariable=self.subject_var)
        self.load_subjects()
        self.subject_combo.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Дата урока (ДД.ММ.ГГГГ)").pack(pady=5)
        self.lesson_date_entry = tk.Entry(self)
        self.lesson_date_entry.pack(fill=tk.X, padx=10)
        
        tk.Label(self, text="Описание задания").pack(pady=5)
        self.description_text = tk.Text(self, height=10)
        self.description_text.pack(fill=tk.BOTH, padx=10, pady=5)
        
        tk.Label(self, text="Срок выполнения (ДД.ММ.ГГГГ)").pack(pady=5)
        self.due_date_entry = tk.Entry(self)
        self.due_date_entry.pack(fill=tk.X, padx=10)
        
        tk.Button(self, text="Сохранить", command=self.save_homework).pack(pady=20)
    
    def load_classes(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT class_name FROM classes")
        classes = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.class_combo['values'] = classes
    
    def load_subjects(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("SELECT subject_name FROM subjects")
        subjects = [row[0] for row in cursor.fetchall()]
        conn.close()
        self.subject_combo['values'] = subjects
    
    def load_homework_data(self):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            SELECT classes.class_name, subjects.subject_name, homework.lesson_date, homework.description, homework.due_date
            FROM homework
            LEFT JOIN classes ON homework.class_id = classes.id
            LEFT JOIN subjects ON homework.subject_id = subjects.id
            WHERE homework.id=?
        """, (self.homework_id,))
        homework = cursor.fetchone()
        conn.close()
        if homework:
            self.class_var.set(homework[0])
            self.subject_var.set(homework[1])
            self.lesson_date_entry.insert(0, homework[2])
            self.description_text.insert(tk.END, homework[3])
            self.due_date_entry.insert(0, homework[4])
    
    def save_homework(self):
        cls = self.class_var.get()
        subject = self.subject_var.get()
        lesson_date = self.lesson_date_entry.get()
        description = self.description_text.get("1.0", tk.END).strip()
        due_date = self.due_date_entry.get()
        
        if not all([cls, subject, lesson_date, description, due_date]):
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все обязательные поля.")
            return
        
        # Валидация дат
        try:
            datetime.strptime(lesson_date, "%d.%m.%Y")
            datetime.strptime(due_date, "%d.%m.%Y")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат даты.")
            return
        
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        
        # Получение id класса
        cursor.execute("SELECT id FROM classes WHERE class_name=?", (cls,))
        class_row = cursor.fetchone()
        if not class_row:
            messagebox.showerror("Ошибка", "Класс не найден.")
            conn.close()
            return
        class_id = class_row[0]
        
        # Получение id предмета
        cursor.execute("SELECT id FROM subjects WHERE subject_name=?", (subject,))
        subject_row = cursor.fetchone()
        if not subject_row:
            messagebox.showerror("Ошибка", "Предмет не найден.")
            conn.close()
            return
        subject_id = subject_row[0]
        
        if self.homework_id:
            cursor.execute("""
                UPDATE homework
                SET class_id=?, subject_id=?, lesson_date=?, description=?, due_date=?
                WHERE id=?
            """, (class_id, subject_id, lesson_date, description, due_date, self.homework_id))
            # Обновление события в календаре
            self.update_calendar_event('homework', self.homework_id, title=f"Домашнее задание по {subject}", 
                                       start_time=datetime.strptime(due_date, "%d.%m.%Y"), 
                                       end_time=datetime.strptime(due_date, "%d.%m.%Y") + timedelta(hours=1))
        else:
            cursor.execute("""
                INSERT INTO homework (class_id, subject_id, lesson_date, description, due_date)
                VALUES (?, ?, ?, ?, ?)
            """, (class_id, subject_id, lesson_date, description, due_date))
            homework_id = cursor.lastrowid
            # Добавление события в календарь
            self.add_calendar_event('homework', homework_id, title=f"Домашнее задание по {subject}", 
                                    start_time=datetime.strptime(due_date, "%d.%m.%Y"), 
                                    end_time=datetime.strptime(due_date, "%d.%m.%Y") + timedelta(hours=1))
        
        conn.commit()
        conn.close()
        
        self.refresh_callback()
        messagebox.showinfo("Успех", "Домашнее задание сохранено успешно.")
        self.destroy()

    def add_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO events (title, description, start_time, end_time, status, comments)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (title, '', start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments))
        conn.commit()
        conn.close()

    def update_calendar_event(self, event_type, reference_id, title, start_time, end_time, status='planned', comments=''):
        conn = sqlite3.connect('class_journal.db')
        cursor = conn.cursor()
        # Предположим, что связь между домашним заданием и событием происходит через title и date
        cursor.execute("""
            UPDATE events
            SET title=?, start_time=?, end_time=?, status=?, comments=?
            WHERE title LIKE ? AND start_time LIKE ?
        """, (title, start_time.strftime("%Y-%m-%d %H:%M"), end_time.strftime("%Y-%m-%d %H:%M"), status, comments, f"Домашнее задание по {title.split(' по ')[1]}%", start_time.strftime("%Y-%m-%d %H:%M")))
        conn.commit()
        conn.close()

# --- Класс Замечания ---
class RemarkForm(tk.Toplevel):
    # Определение класса уже приведено выше
    pass

# --- Класс Домашнего Задания ---
class HomeworkForm(tk.Toplevel):
    # Определение класса уже приведено выше
    pass

# --- Класс Отправки Выполнения ---
class SubmissionForm(tk.Toplevel):
    # Определение класса уже приведено выше
    pass

# --- Основное приложение ---
class MainApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Электронный классный журнал")
        self.master.geometry("800x600")
        
        self.create_menu()
        self.create_main_frame()
    
    def create_menu(self):
        menubar = tk.Menu(self.master)
        
        # Лист всеобуча
        menubar.add_command(label="Лист всеобуча", command=self.open_student_list)
        
        # Оценки
        menubar.add_command(label="Оценки", command=self.open_grades)
        
        # Пропуски
        menubar.add_command(label="Пропуски", command=self.open_absences)
        
        # Замечания
        menubar.add_command(label="Замечания", command=self.open_remarks)
        
        # Домашнее задание
        menubar.add_command(label="Домашнее задание", command=self.open_homework)
        
        # Печать
        menubar.add_command(label="Печать", command=self.print_journal)
        
        # Календарь
        menubar.add_command(label="Календарь", command=self.open_calendar)
        
        # Профиль
        menubar.add_command(label="Профиль", command=self.open_profile)
        
        # Выход
        menubar.add_command(label="Выход", command=self.master.quit)
        
        self.master.config(menu=menubar)
    
    def create_main_frame(self):
        self.main_frame = tk.Frame(self.master)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        welcome_label = tk.Label(self.main_frame, text="Добро пожаловать в Электронный классный журнал!", font=("Arial", 16))
        welcome_label.pack(pady=20)
    
    def open_student_list(self):
        StudentListWindow(self.master)
    
    def open_grades(self):
        GradesWindow(self.master)
    
    def open_absences(self):
        AbsencesWindow(self.master)
    
    def open_remarks(self):
        RemarksWindow(self.master)
    
    def open_homework(self):
        HomeworkWindow(self.master)
    
    def print_journal(self):
        PrintJournalWindow(self.master)
    
    def open_calendar(self):
        CalendarApp(self.master)
    
    def open_profile(self):
        ProfileWindow(self.master)

# --- Запуск приложения ---
if __name__ == "__main__":
    initialize_db()
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
